"""Fill Huawei Individual Developers application xlsx for Personal Agent."""
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

SRC = Path(r"c:\Users\Ababikov\Downloads\Application+Material+(for+Individual+Developers).xlsx")
OUT = Path(__file__).resolve().parents[1] / "docs" / "huawei" / "Application-Material-Individual-PersonalAgent-filled.xlsx"

BLACK = Font(color="000000", name="Calibri", size=11)


def set_cell(ws, coord: str, value: str) -> None:
    cell = ws[coord]
    cell.value = value
    cell.font = BLACK


def clear_row(ws, row: int, cols: str = "ABCDEFGHI") -> None:
    for col in cols:
        ws[f"{col}{row}"].value = None


def main() -> None:
    wb = load_workbook(SRC)
    ws = wb["Data Usage (Mandatory)"]

    # Row 3 — activity records
    set_cell(ws, "A3", "1")
    set_cell(ws, "B3", "Read user activity record data")
    set_cell(
        ws,
        "C3",
        "Import workout sessions from Huawei Health: start/end time, activity type, "
        "duration, and calories from the device into a personal workout journal "
        "(gym and swimming).",
    )
    set_cell(
        ws,
        "D3",
        "After the user taps Connect in Profile and grants OAuth consent, the app syncs "
        "workout sessions (start/end time, duration, calories, heart rate when available) "
        "and stores them in the database. User can link a device session to a manual "
        "workout entry to compare device calories vs MET estimate.",
    )
    set_cell(ws, "E3", "Not required")
    # F3:F7 merged — client type once for all scopes
    set_cell(ws, "F3", "Website (H5)")
    set_cell(ws, "G3", "2026-06-30")

    # Row 4 — historical data (max period Huawei grants, typically 1 year)
    set_cell(ws, "A4", "2")
    set_cell(ws, "B4", "Read historical data (1 year)")
    set_cell(
        ws,
        "C4",
        "Read activity records from up to one year before the user's first authorization "
        "(OAuth scope: history data open). Required to import the user's full workout "
        "history from Huawei Health into the personal journal, not only new sessions "
        "after connect.",
    )
    set_cell(
        ws,
        "D4",
        "On first connect, backfill all gym and swim sessions stored in Huawei Health "
        "within the approved historical window. User explicitly opts in on the Huawei "
        "authorization screen. Single-user personal app.",
    )
    set_cell(ws, "E4", "Not required")
    set_cell(ws, "G4", "2026-06-30")

    # Row 5 — heart rate (matches permission form)
    set_cell(ws, "A5", "3")
    set_cell(ws, "B5", "Read heart rate data")
    set_cell(
        ws,
        "C5",
        "Read heart rate associated with workout activity records (average HR when "
        "provided in activity summary from Huawei Health).",
    )
    set_cell(
        ws,
        "D5",
        "Display average heart rate from the watch/phone next to imported workout "
        "sessions in the personal journal.",
    )
    set_cell(ws, "E5", "Not required")
    set_cell(ws, "G5", "2026-06-30")

    for r in range(6, 12):
        clear_row(ws, r, cols="ABCDE")

    # App Info
    ai = wb["App Info (Mandatory)"]
    set_cell(
        ai,
        "B2",
        "Personal Agent is a private web app for tracking gym and swim workouts, "
        "calories (MET estimate), and a progress calendar. Optional Huawei Health "
        "integration imports activity records after explicit user consent. "
        "URL: https://personal-agent-zeta.vercel.app",
    )
    set_cell(
        ai,
        "B3",
        "Screenshots: (1) workout calendar home, (2) Profile with Huawei Health "
        "Connect/Sync, (3) privacy page. Attach JPG/PNG in Huawei portal if required.",
    )
    set_cell(
        ai,
        "B4",
        "App type: Web application (HTML5). Industry: Health & Fitness / Personal wellness.",
    )
    set_cell(
        ai,
        "B5",
        "Personal project for the developer's own training log; not distributed to the public.",
    )
    set_cell(
        ai,
        "B6",
        "Activity record read + 1-year historical data scopes are used to import the "
        "user's full workout history from Huawei Health and link sessions to manual "
        "journal entries without creating duplicate workouts.",
    )

    # Self-Check
    sc = wb["Self-Check (Mandatory)"]
    set_cell(sc, "C2", "Yes. The app is available in Russia. URL: https://personal-agent-zeta.vercel.app")
    set_cell(sc, "C3", "Demo (personal use, test phase).")
    set_cell(sc, "C4", "No.")
    set_cell(sc, "C5", "Website (HTML5 web application).")
    set_cell(sc, "C6", "RESTful API.")
    set_cell(sc, "C7", "Yes.")
    set_cell(sc, "C8", "No.")
    set_cell(sc, "C9", "1 user (personal deployment), TPS < 10.")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    main()
